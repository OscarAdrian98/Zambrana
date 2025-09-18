from flask import render_template, redirect, url_for, request, flash, send_file, make_response, session
import io
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from app import app, db, login_manager
from app.models import Usuario, Fichaje, Modificacion, RegistroHorario, Ausencia, SabadoAsignado
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl.styles import Font, Alignment
import logging
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import pytz
from itertools import groupby

# Zona horaria de España
zona_es = pytz.timezone("Europe/Madrid")

# Crear carpeta logs si no existe
os.makedirs('logs', exist_ok=True)

# Configurar logging
logging.basicConfig(filename='logs/flask_app.log', level=logging.INFO,
                    format='%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')

@app.context_processor
def inject_today():
    return {'today': date.today()}

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        remember = True if request.form.get('remember') else False  # 👈 Nuevo

        user = Usuario.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=remember)  # 👈 Nuevo parámetro
            return redirect(url_for('fichar'))
        else:
            flash("Usuario o contraseña incorrectos")

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/fichar', methods=['GET', 'POST'])
@login_required
def fichar():
    today = date.today()
    fichaje = Fichaje.query.filter_by(usuario_id=current_user.id, fecha=today).first()

    total_segundos = 0
    ultima_entrada_str = None  # NUEVO

    if fichaje and fichaje.registros:
        registros_hoy = sorted([r for r in fichaje.registros if r.timestamp.date() == today], key=lambda r: r.timestamp)

        # Calcular horas completas en pares entrada/salida
        for i in range(0, len(registros_hoy) - 1, 2):
            entrada = registros_hoy[i]
            salida = registros_hoy[i + 1]
            if entrada.tipo == 'entrada' and salida.tipo == 'salida':
                diff = (salida.timestamp - entrada.timestamp).total_seconds()
                if diff > 0:
                    total_segundos += diff

        # Si hay un registro impar y termina en 'entrada', guardar su timestamp
        if len(registros_hoy) % 2 == 1 and registros_hoy[-1].tipo == 'entrada':
            ultima_entrada_str = registros_hoy[-1].timestamp.isoformat()

    horas = int(total_segundos // 3600)
    minutos = int((total_segundos % 3600) // 60)

    # NUEVO: comprobar si hoy es sábado y está asignado
    es_sabado = today.weekday() == 5 and today.month != 8
    sabado_asignado = SabadoAsignado.query.filter_by(usuario_id=current_user.id, fecha=today).first() is not None

    return render_template(
        "fichar.html",
        fichaje=fichaje,
        today=today,
        horas=horas,
        minutos=minutos,
        es_sabado=es_sabado,
        sabado_asignado=sabado_asignado,
        ultima_entrada_str=ultima_entrada_str
    )


@app.route('/fichar_registro', methods=['POST'])
@login_required
def fichar_registro():
    tipo = request.form.get('tipo')  # 'entrada' o 'salida'
    origen = request.form.get('origen', 'Tienda')  # Captura origen del formulario

    now_es = datetime.now(zona_es)
    hoy = now_es.date()

    # ❌ Bloquear si el usuario tiene una ausencia hoy
    tiene_ausencia = Ausencia.query.filter_by(usuario_id=current_user.id, fecha=hoy).first()
    if tiene_ausencia:
        flash(f"No puedes fichar hoy porque estás ausente ({tiene_ausencia.tipo}).", "danger")
        return redirect(url_for('fichar'))

    if tipo not in ['entrada', 'salida']:
        flash("Tipo de fichaje no válido.")
        return redirect(url_for('fichar'))

    try:
        # Buscar o crear el fichaje del día
        fichaje = Fichaje.query.filter_by(usuario_id=current_user.id, fecha=hoy).first()
        if not fichaje:
            if tipo == 'salida':
                flash("No puedes registrar una salida sin haber fichado una entrada antes.")
                return redirect(url_for('fichar'))

            # Crear fichaje si es una entrada
            fichaje = Fichaje(
                usuario_id=current_user.id,
                fecha=hoy,
                fecha_creacion=now_es,
                creado_por_admin=current_user.admin,
                eliminado=False
            )
            db.session.add(fichaje)
            db.session.commit()

        # Obtener los registros del día ordenados
        registros = RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False)\
            .order_by(RegistroHorario.timestamp).all()

        if registros:
            ultimo = registros[-1]
            if ultimo.tipo == tipo:
                flash(f"Ya has registrado una {tipo} como último fichaje. Debes registrar una {'salida' if tipo == 'entrada' else 'entrada'} antes.")
                return redirect(url_for('fichar'))

        elif tipo == 'salida':
            flash("No puedes registrar una salida sin haber fichado una entrada antes.")
            return redirect(url_for('fichar'))

        # Crear nuevo registro válido con origen
        nuevo_registro = RegistroHorario(
            fichaje_id=fichaje.id,
            tipo=tipo,
            timestamp=now_es,
            creado_por_admin=current_user.admin,
            eliminado=False,
            origen=origen
        )
        db.session.add(nuevo_registro)
        db.session.commit()

        flash(f"{tipo.capitalize()} registrada correctamente a las {now_es.strftime('%H:%M:%S')} ({origen}).")

    except Exception as e:
        logging.error("Error al registrar fichaje", exc_info=True)
        flash("Error al registrar el fichaje.")

    return redirect(url_for('fichar'))

@app.route('/resumen')
@login_required
def resumen():
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    hoy = date.today()
    
    # Si no hay filtros, usamos el primer y último día del mes actual
    if not desde_str:
        desde = hoy.replace(day=1)
        desde_str = desde.strftime('%Y-%m-%d')
    else:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()

    if not hasta_str:
        # Para obtener el último día del mes, avanzamos al día 1 del mes siguiente y restamos un día
        if hoy.month == 12:
            siguiente_mes = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            siguiente_mes = hoy.replace(month=hoy.month + 1, day=1)
        hasta = siguiente_mes - timedelta(days=1)
        hasta_str = hasta.strftime('%Y-%m-%d')
    else:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()

    query = Fichaje.query.filter_by(usuario_id=current_user.id).filter(
        Fichaje.eliminado == False,
        Fichaje.fecha >= desde,
        Fichaje.fecha <= hasta
    )

    fichajes = query.order_by(Fichaje.fecha.desc()).all()

    datos = []

    for f in fichajes:
        registros = (
            RegistroHorario.query
            .filter_by(fichaje_id=f.id, eliminado=False)
            .order_by(RegistroHorario.timestamp)
            .all()
        )

        registros_simplificados = []
        reg_buffer = None
        total = timedelta()

        for reg in registros:
            hora_str = reg.timestamp.strftime('%H:%M:%S')
            registros_simplificados.append({
                'tipo': reg.tipo,
                'hora': hora_str,
                'origen': reg.origen or 'Tienda'
            })

            if reg.tipo == 'entrada':
                reg_buffer = reg.timestamp
            elif reg.tipo == 'salida' and reg_buffer:
                diff = reg.timestamp - reg_buffer
                if diff.total_seconds() > 0:
                    total += diff
                reg_buffer = None

        datos.append({
            'fecha': f.fecha.strftime('%d/%m/%Y'),
            'registros': registros_simplificados,
            'total': str(total) if total.total_seconds() > 0 else ''
        })

    return render_template('resumen.html', datos=datos, desde=desde_str, hasta=hasta_str)

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    usuarios = Usuario.query.order_by(Usuario.nombre).all()
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/fichajes')
@login_required
def admin_fichajes():
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    hoy = date.today()

    # Rango por defecto: mes actual
    if not desde_str:
        desde = hoy.replace(day=1)
        desde_str = desde.strftime('%Y-%m-%d')
    else:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()

    if not hasta_str:
        if hoy.month == 12:
            siguiente_mes = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            siguiente_mes = hoy.replace(month=hoy.month + 1, day=1)
        hasta = siguiente_mes - timedelta(days=1)
        hasta_str = hasta.strftime('%Y-%m-%d')
    else:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()

    query = Fichaje.query.join(Usuario).filter(
        Fichaje.eliminado == False,
        Fichaje.fecha.between(desde, hasta)
    )

    fichajes = query.order_by(Fichaje.fecha.desc()).all()

    datos = []
    for f in fichajes:
        # Filtrar registros válidos y ordenarlos
        registros = sorted(
            [r for r in f.registros if not r.eliminado],
            key=lambda r: r.timestamp
        )

        total = timedelta()
        for i in range(0, len(registros), 2):
            entrada = registros[i].timestamp if i < len(registros) else None
            salida = registros[i + 1].timestamp if i + 1 < len(registros) else None
            if entrada and salida and entrada < salida:
                total += salida - entrada

        datos.append({
            'usuario_id': f.usuario.id,
            'usuario': f.usuario.nombre,
            'fecha': f.fecha.strftime('%d/%m/%Y'),
            'total': str(total) if total.total_seconds() > 0 else '',
            'fichaje_id': f.id
        })

    return render_template(
        'admin_fichajes.html',
        datos=datos,
        desde_str=desde_str,
        hasta_str=hasta_str
    )

@app.route('/admin/exportar_excel')
@login_required
def exportar_excel():
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    query = Fichaje.query.join(Usuario).filter(Fichaje.eliminado == False)

    if desde_str:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha >= desde)
    if hasta_str:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha <= hasta)

    fichajes = query.order_by(Fichaje.fecha.desc()).all()

    data = []
    for f in fichajes:
        registros = RegistroHorario.query.filter_by(fichaje_id=f.id, eliminado=False)\
            .order_by(RegistroHorario.timestamp).all()
        
        # Agrupar en pares de tramos: [(entrada1, salida1), (entrada2, salida2), ...]
        tramos = [
            (registros[i], registros[i+1])
            for i in range(0, len(registros), 2)
            if i + 1 < len(registros)
        ]

        # Mañana = primer tramo, Tarde = segundo tramo (si existen)
        entrada_m = tramos[0][0].timestamp.time() if len(tramos) > 0 else None
        salida_m  = tramos[0][1].timestamp.time() if len(tramos) > 0 else None
        entrada_t = tramos[1][0].timestamp.time() if len(tramos) > 1 else None
        salida_t  = tramos[1][1].timestamp.time() if len(tramos) > 1 else None

        total = timedelta()
        if entrada_m and salida_m:
            total += datetime.combine(f.fecha, salida_m) - datetime.combine(f.fecha, entrada_m)
        if entrada_t and salida_t:
            total += datetime.combine(f.fecha, salida_t) - datetime.combine(f.fecha, entrada_t)

        data.append({
            'Usuario': f.usuario.nombre,
            'Fecha': f.fecha.strftime('%d/%m/%Y'),
            'Entrada Mañana': entrada_m.strftime('%H:%M') if entrada_m else '',
            'Salida Mañana': salida_m.strftime('%H:%M') if salida_m else '',
            'Entrada Tarde': entrada_t.strftime('%H:%M') if entrada_t else '',
            'Salida Tarde': salida_t.strftime('%H:%M') if salida_t else '',
            'Total trabajado': str(total) if total.total_seconds() > 0 else ''
        })

    # Generar Excel
    df = pd.DataFrame(data)
    output = io.BytesIO()
    filename = f"fichajes_{desde_str or 'inicio'}_a_{hasta_str or 'hoy'}.xlsx"

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fichajes')

    output.seek(0)
    return send_file(output, download_name=filename, as_attachment=True)

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        puesto = request.form['puesto']

        existente = Usuario.query.filter_by(email=email).first()
        if existente:
            flash("Ya existe un usuario con ese email.")
            return redirect(url_for('registro'))

        password_hash = generate_password_hash(password)
        nuevo_usuario = Usuario(
            nombre=nombre,
            email=email,
            password_hash=password_hash,
            puesto=puesto,
            admin=False
        )
        db.session.add(nuevo_usuario)
        db.session.commit()

        login_user(nuevo_usuario)
        flash("Registro exitoso. Bienvenido.")
        return redirect(url_for('fichar'))

    return render_template('registro.html')

@app.route('/exportar_mis_fichajes')
@login_required
def exportar_mis_fichajes():
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    query = Fichaje.query.filter_by(usuario_id=current_user.id).filter(Fichaje.eliminado == False)

    if desde_str:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha >= desde)

    if hasta_str:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha <= hasta)

    fichajes = query.order_by(Fichaje.fecha).all()

    data = []
    total_general = timedelta()

    for f in fichajes:
        registros = (
            RegistroHorario.query
            .filter_by(fichaje_id=f.id, eliminado=False)
            .order_by(RegistroHorario.timestamp)
            .all()
        )

        lista_registros = []
        reg_buffer = None
        total_dia = timedelta()

        for r in registros:
            hora_str = r.timestamp.strftime('%H:%M:%S')
            origen = r.origen or 'Tienda'
            lista_registros.append(f"{r.tipo.capitalize()} - {hora_str} ({origen})")

            if r.tipo == 'entrada':
                reg_buffer = r.timestamp
            elif r.tipo == 'salida' and reg_buffer:
                diff = r.timestamp - reg_buffer
                if diff.total_seconds() > 0:
                    total_dia += diff
                reg_buffer = None

        total_general += total_dia

        data.append({
            'Fecha': f.fecha.strftime('%d/%m/%Y'),
            'Registros': "\n".join(lista_registros) if lista_registros else '-',
            'Total trabajado': str(total_dia) if total_dia.total_seconds() > 0 else '-'
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    filename = f"fichajes_{current_user.nombre.replace(' ', '_')}.xlsx"

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fichajes', startrow=2)

        workbook = writer.book
        worksheet = writer.sheets['Fichajes']

        titulo = f"Horario de {current_user.nombre}"
        worksheet.merge_cells('A1:C1')
        cell = worksheet['A1']
        cell.value = titulo
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')

        total_filas = len(df) + 4
        worksheet.cell(row=total_filas, column=1, value="Total acumulado:")
        worksheet.cell(row=total_filas, column=3, value=str(total_general))

    output.seek(0)
    return send_file(output, download_name=filename, as_attachment=True)

@app.route('/admin/fichajes/<int:usuario_id>')
@login_required
def admin_fichajes_usuario(usuario_id):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    usuario = Usuario.query.get_or_404(usuario_id)

    # Obtener fechas de filtro o establecer por defecto el mes actual
    hoy = date.today()
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    if desde_str:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
    else:
        desde = hoy.replace(day=1)
        desde_str = desde.strftime('%Y-%m-%d')

    if hasta_str:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
    else:
        if hoy.month == 12:
            siguiente_mes = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            siguiente_mes = hoy.replace(month=hoy.month + 1, day=1)
        hasta = siguiente_mes - timedelta(days=1)
        hasta_str = hasta.strftime('%Y-%m-%d')

    fichajes = Fichaje.query.filter_by(usuario_id=usuario_id, eliminado=False).filter(
        Fichaje.fecha.between(desde, hasta)
    ).order_by(Fichaje.fecha.desc()).all()

    datos = []
    for f in fichajes:
        registros = (
            RegistroHorario.query
            .filter_by(fichaje_id=f.id, eliminado=False)
            .order_by(RegistroHorario.timestamp)
            .all()
        )

        lista_registros = []
        reg_buffer = None
        total_dia = timedelta()

        for r in registros:
            hora_str = r.timestamp.strftime('%H:%M:%S')
            origen = r.origen or 'Tienda'
            lista_registros.append(f"{r.tipo.capitalize()} - {hora_str} ({origen})")

            if r.tipo == 'entrada':
                reg_buffer = r.timestamp
            elif r.tipo == 'salida' and reg_buffer:
                diff = r.timestamp - reg_buffer
                if diff.total_seconds() > 0:
                    total_dia += diff
                reg_buffer = None

        datos.append({
            'id': f.id,
            'fecha': f.fecha.strftime('%d/%m/%Y'),
            'registros': lista_registros,
            'total': str(total_dia) if total_dia.total_seconds() > 0 else ''
        })

    return render_template('admin_fichajes_usuario.html',
                           usuario=usuario,
                           datos=datos,
                           desde_str=desde_str,
                           hasta_str=hasta_str)

@app.route('/admin/editar_fichaje/<int:fichaje_id>', methods=['GET', 'POST'])
@login_required
def editar_fichaje_admin(fichaje_id):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    fichaje = Fichaje.query.get_or_404(fichaje_id)

    if request.method == 'POST':
        try:
            entradas = request.form.getlist('entradas[]')
            salidas = request.form.getlist('salidas[]')
            registro_ids = request.form.getlist('registro_ids[]')

            # Marcar todos como eliminados por defecto
            RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False).update({'eliminado': True})
            db.session.flush()

            tramos_guardados = 0

            for i in range(len(entradas)):
                hora_entrada_str = entradas[i].strip()
                hora_salida_str = salidas[i].strip()

                if not hora_entrada_str or not hora_salida_str:
                    continue

                hora_entrada = datetime.strptime(hora_entrada_str, '%H:%M')
                hora_salida = datetime.strptime(hora_salida_str, '%H:%M')

                if hora_entrada >= hora_salida:
                    flash(f"La entrada debe ser anterior a la salida (tramo {i + 1}).")
                    return redirect(request.url)

                # Comprobar si existían registros previos
                try:
                    id_entrada = int(registro_ids[i * 2])
                    id_salida = int(registro_ids[i * 2 + 1])

                    r_entrada_ant = RegistroHorario.query.get(id_entrada)
                    r_salida_ant = RegistroHorario.query.get(id_salida)

                    if r_entrada_ant and r_salida_ant:
                        misma_entrada = r_entrada_ant.timestamp.strftime('%H:%M') == hora_entrada_str
                        misma_salida = r_salida_ant.timestamp.strftime('%H:%M') == hora_salida_str

                        if misma_entrada and misma_salida:
                            # Restaurar los originales si no han cambiado
                            r_entrada_ant.eliminado = False
                            r_salida_ant.eliminado = False
                            continue

                except (IndexError, ValueError):
                    pass  # Tramo nuevo

                # Guardar nuevos registros porque han cambiado o son nuevos
                r1 = RegistroHorario(
                    fichaje_id=fichaje.id,
                    tipo='entrada',
                    timestamp=datetime.combine(fichaje.fecha, hora_entrada.time()),
                    creado_por_admin=True,
                    eliminado=False
                )
                r2 = RegistroHorario(
                    fichaje_id=fichaje.id,
                    tipo='salida',
                    timestamp=datetime.combine(fichaje.fecha, hora_salida.time()),
                    creado_por_admin=True,
                    eliminado=False
                )
                db.session.add_all([r1, r2])

                mod = Modificacion(
                    fichaje_id=fichaje.id,
                    admin_id=current_user.id,
                    campo_modificado=f"tramo_{i + 1}",
                    valor_anterior="-",
                    valor_nuevo=f"{hora_entrada_str} → {hora_salida_str}",
                    fecha_modificacion=datetime.now()
                )
                db.session.add(mod)

                tramos_guardados += 1

            if tramos_guardados == 0:
                registros_previos = RegistroHorario.query.filter_by(fichaje_id=fichaje.id).order_by(RegistroHorario.timestamp).all()

                tramos_previos = []
                for i in range(0, len(registros_previos), 2):
                    entrada = registros_previos[i].timestamp.strftime('%H:%M') if i < len(registros_previos) else ''
                    salida = registros_previos[i + 1].timestamp.strftime('%H:%M') if i + 1 < len(registros_previos) else ''
                    if entrada and salida:
                        tramos_previos.append(f"{entrada} → {salida}")

                # SOLO último tramo
                valor_anterior = tramos_previos[-1] if tramos_previos else "-"
                valor_nuevo = "Eliminados todos los tramos o no hubo cambios"

                mod = Modificacion(
                    fichaje_id=fichaje.id,
                    admin_id=current_user.id,
                    campo_modificado="fichaje",
                    valor_anterior=valor_anterior,
                    valor_nuevo=valor_nuevo,
                    fecha_modificacion=datetime.now()
                )
                db.session.add(mod)

            db.session.commit()
            flash("Fichaje actualizado correctamente.")
            return redirect(url_for('admin_fichajes_usuario', usuario_id=fichaje.usuario_id))

        except Exception as e:
            import logging
            logging.exception("Error al editar fichaje")
            flash("Error procesando la edición del fichaje.")
            return redirect(request.url)

    registros_visibles = RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False)\
        .order_by(RegistroHorario.timestamp).all()

    print(f"Registros encontrados para fichaje {fichaje.id}: {len(registros_visibles)}")
    for r in registros_visibles:
        print(f" - {r.tipo} a las {r.timestamp}")

    return render_template(
        'editar_fichaje_admin.html',
        fichaje=fichaje,
        registros_visibles=registros_visibles
    )

@app.route('/admin/borrar_fichaje/<int:fichaje_id>')
@login_required
def borrar_fichaje_admin(fichaje_id):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    fichaje = Fichaje.query.get_or_404(fichaje_id)
    usuario_id = fichaje.usuario_id
    fichaje.eliminado = True
    db.session.commit()
    flash("Fichaje eliminado correctamente.")
    return redirect(url_for('admin_fichajes_usuario', usuario_id=usuario_id))

@app.route('/exportar_mis_fichajes_pdf')
@login_required
def exportar_mis_fichajes_pdf():
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')

    query = Fichaje.query.filter_by(usuario_id=current_user.id).filter(Fichaje.eliminado == False)

    if desde_str:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha >= desde)

    if hasta_str:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        query = query.filter(Fichaje.fecha <= hasta)

    fichajes = query.order_by(Fichaje.fecha).all()

    data = [["Fecha", "Registros", "Total trabajado"]]
    total_general = timedelta()

    for f in fichajes:
        registros = RegistroHorario.query.filter_by(fichaje_id=f.id, eliminado=False).order_by(RegistroHorario.timestamp).all()

        lista_registros = []
        reg_buffer = None
        total_dia = timedelta()

        for r in registros:
            hora_str = r.timestamp.astimezone(zona_es).strftime('%H:%M:%S')
            origen = r.origen or 'Tienda'
            lista_registros.append(f"{r.tipo.capitalize()} - {hora_str} ({origen})")

            if r.tipo == 'entrada':
                reg_buffer = r.timestamp
            elif r.tipo == 'salida' and reg_buffer:
                diff = r.timestamp - reg_buffer
                if diff.total_seconds() > 0:
                    total_dia += diff
                reg_buffer = None

        total_general += total_dia

        data.append([
            f.fecha.strftime('%d/%m/%Y'),
            "\n".join(lista_registros) if lista_registros else "-",
            str(total_dia) if total_dia.total_seconds() > 0 else "-"
        ])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=4*cm, bottomMargin=3*cm)

    styles = getSampleStyleSheet()
    elementos = []

    # Tabla de fichajes
    tabla = Table(data, colWidths=[3*cm, 8.5*cm, 4.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.5*cm))

    # Total acumulado
    total_horas = int(total_general.total_seconds() // 3600)
    total_minutos = int((total_general.total_seconds() % 3600) // 60)
    total_segundos = int(total_general.total_seconds() % 60)
    total_formateado = f"{total_horas:02}:{total_minutos:02}:{total_segundos:02}"

    p_total = Paragraph(f"<b>Total acumulado:</b> {total_formateado}", styles['Normal'])
    elementos.append(p_total)

    # Encabezado y pie para cada página
    def encabezado(canvas, doc):
        width, height = A4
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo-zambrana.jpg')

        if os.path.exists(logo_path):
            try:
                canvas.drawImage(logo_path, 2*cm, height - 3*cm, width=6*cm, preserveAspectRatio=True, mask='auto')
            except Exception as e:
                pass  # Evitar que un error del logo rompa todo

        canvas.setFont("Helvetica-Bold", 22)
        canvas.drawCentredString(width / 2, height - 2.8*cm, "FICHAJES")

        canvas.setFont("Helvetica", 13)
        partes = []
        if desde_str:
            partes.append(f"desde {datetime.strptime(desde_str, '%Y-%m-%d').strftime('%d/%m/%Y')}")
        if hasta_str:
            partes.append(f"hasta {datetime.strptime(hasta_str, '%Y-%m-%d').strftime('%d/%m/%Y')}")
        subtitulo = f"{' '.join(partes)} - {current_user.nombre}"
        canvas.drawCentredString(width / 2, height - 3.5*cm, subtitulo)

        canvas.line(2*cm, height - 3.9*cm, width - 2*cm, height - 3.9*cm)

        # Pie
        fecha_actual = datetime.now(zona_es).strftime('%d/%m/%Y')
        canvas.setFont("Helvetica-Oblique", 8)
        canvas.drawRightString(width - 2*cm, 1.5*cm, f"Documento generado el {fecha_actual}")
        canvas.drawString(2*cm, 1.5*cm, "Horario expresado en hora local (Europe/Madrid)")

    # Construcción final
    doc.build(elementos, onFirstPage=encabezado, onLaterPages=encabezado)

    buffer.seek(0)
    filename = f"fichajes_{current_user.nombre.replace(' ', '_')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

# Función auxiliar para formatear fechas en español
def formatear_fecha_es(fecha):
    meses = [
        'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
        'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'
    ]
    return f"{fecha.day} de {meses[fecha.month - 1]} de {fecha.year}"

@app.route('/mis-ausencias', methods=['GET', 'POST'])
@login_required
def mis_ausencias():
    if request.method == 'POST':
        desde_str = request.form.get('fecha_desde')
        hasta_str = request.form.get('fecha_hasta')
        tipo = request.form.get('tipo')
        observaciones = request.form.get('observaciones', '').strip()

        # --- Si es edición, eliminar el tramo anterior ---
        editar_fecha_inicio = request.form.get('editar_fecha_inicio')
        editar_fecha_fin = request.form.get('editar_fecha_fin')
        editar_tipo = request.form.get('editar_tipo')
        editar_observaciones = request.form.get('editar_observaciones')

        if editar_fecha_inicio and editar_fecha_fin and editar_tipo is not None:
            try:
                f_ini = datetime.strptime(editar_fecha_inicio, '%Y-%m-%d').date()
                f_fin = datetime.strptime(editar_fecha_fin, '%Y-%m-%d').date()
                ausencias_antiguas = Ausencia.query.filter(
                    Ausencia.usuario_id == current_user.id,
                    Ausencia.fecha >= f_ini,
                    Ausencia.fecha <= f_fin,
                    Ausencia.tipo == editar_tipo,
                    (Ausencia.observaciones == editar_observaciones) |
                    ((Ausencia.observaciones.is_(None)) if not editar_observaciones else False)
                ).all()
                for a in ausencias_antiguas:
                    db.session.delete(a)
                db.session.commit()
            except Exception as e:
                flash("Error al eliminar la ausencia antigua.", "danger")

        # --- Crear la nueva ausencia (insert normal) ---
        try:
            fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()

            if fecha_hasta < fecha_desde:
                flash('La fecha "Hasta" no puede ser anterior a "Desde".', 'danger')
                return redirect(url_for('mis_ausencias'))

            dias_ausencia = [
                fecha_desde + timedelta(days=i)
                for i in range((fecha_hasta - fecha_desde).days + 1)
            ]

            creadas = 0
            for dia in dias_ausencia:
                ya_existe = Ausencia.query.filter_by(usuario_id=current_user.id, fecha=dia).first()
                if not ya_existe:
                    nueva = Ausencia(
                        usuario_id=current_user.id,
                        fecha=dia,
                        tipo=tipo,
                        observaciones=observaciones,
                        creado_por_admin=False
                    )
                    db.session.add(nueva)
                    creadas += 1

            db.session.commit()

            if creadas:
                flash(f'{creadas} día(s) de ausencia registrados correctamente.', 'success')
            else:
                flash('No se registraron nuevas ausencias porque ya existían.', 'warning')

            return redirect(url_for('mis_ausencias'))

        except ValueError:
            flash('Fechas inválidas.', 'danger')
            return redirect(url_for('mis_ausencias'))

    # --------------------
    # MODO GET - Cargar ausencias y agrupar por tramos consecutivos
    # --------------------
    hoy = date.today()
    anyo_actual = hoy.year

    ausencias = Ausencia.query.filter_by(usuario_id=current_user.id).filter(
        Ausencia.fecha >= date(anyo_actual, 1, 1),
        Ausencia.fecha <= date(anyo_actual, 12, 31)
    ).order_by(Ausencia.tipo, Ausencia.observaciones, Ausencia.fecha).all()

    ausencias_actuales = []

    for (tipo, observaciones), grupo in groupby(ausencias, key=lambda a: (a.tipo, a.observaciones)):
        grupo = list(grupo)
        if not grupo:
            continue

        bloque = []
        anterior = None
        for a in grupo:
            if not anterior or (a.fecha - anterior).days == 1:
                bloque.append(a)
            else:
                ausencias_actuales.append({
                    'fecha': bloque[0].fecha,
                    'fecha_hasta': bloque[-1].fecha,
                    'fecha_texto': formatear_fecha_es(bloque[0].fecha),
                    'fecha_hasta_texto': formatear_fecha_es(bloque[-1].fecha),
                    'tipo': tipo,
                    'observaciones': observaciones,
                    'estado': 'utilizadas' if bloque[-1].fecha < hoy else 'pendiente',
                    'dias': len(bloque),
                    'creado_por_admin': any(a.creado_por_admin for a in bloque)
                })
                bloque = [a]
            anterior = a.fecha

        # Último tramo
        if bloque:
            ausencias_actuales.append({
                'fecha': bloque[0].fecha,
                'fecha_hasta': bloque[-1].fecha,
                'fecha_texto': formatear_fecha_es(bloque[0].fecha),
                'fecha_hasta_texto': formatear_fecha_es(bloque[-1].fecha),
                'tipo': tipo,
                'observaciones': observaciones,
                'estado': 'utilizadas' if bloque[-1].fecha < hoy else 'pendiente',
                'dias': len(bloque),
                'creado_por_admin': any(a.creado_por_admin for a in bloque)
            })

    # --------------------
    # Calcular saldos
    # --------------------
    total_vacaciones = 30

    utilizadas = [a for a in ausencias_actuales if a['tipo'] == 'Vacaciones' and a['fecha_hasta'] < hoy]
    solicitadas = [a for a in ausencias_actuales if a['tipo'] == 'Vacaciones' and a['fecha'] >= hoy]

    saldo_utilizado = sum(a['dias'] for a in utilizadas)
    saldo_solicitado = sum(a['dias'] for a in solicitadas)
    saldo_disponible = total_vacaciones - saldo_utilizado - saldo_solicitado

    # Detectar si hay edición activa
    editar_tramo = session.pop('editar_tramo', None)

    current_date = date.today()

    return render_template('ausencias.html',
                        ausencias_actuales=ausencias_actuales,
                        saldo_disponible=saldo_disponible,
                        saldo_solicitado=saldo_solicitado,
                        saldo_utilizado=saldo_utilizado,
                        editar_tramo=editar_tramo,
                        current_date=current_date)

# -----------------------------------------
# Eliminar ausencia (usuario)
# -----------------------------------------
@app.route('/eliminar-ausencia', methods=['POST'])
@login_required
def eliminar_ausencia_usuario():
    fecha_inicio = request.form.get('fecha_inicio')
    fecha_fin = request.form.get('fecha_fin')
    tipo = request.form.get('tipo')
    observaciones = request.form.get('observaciones')

    try:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

        ausencias = Ausencia.query.filter(
            Ausencia.usuario_id == current_user.id,
            Ausencia.fecha >= fecha_inicio,
            Ausencia.fecha <= fecha_fin,
            Ausencia.tipo == tipo,
            (Ausencia.observaciones == observaciones) | ((Ausencia.observaciones.is_(None)) if not observaciones else False)
        ).all()

        for a in ausencias:
            db.session.delete(a)

        db.session.commit()
        flash("Ausencia eliminada correctamente.", "success")
    except Exception as e:
        flash("Error al eliminar la ausencia.", "danger")

    return redirect(url_for('mis_ausencias'))

# -----------------------------------------
# Preparar edición de ausencia (usuario)
# -----------------------------------------
@app.route('/editar-ausencia', methods=['POST'])
@login_required
def editar_ausencia_usuario():
    # Guardamos los datos del tramo antiguo en sesión para usarlos como edición
    session['editar_tramo'] = {
        'fecha_inicio': request.form.get('fecha_inicio'),
        'fecha_fin': request.form.get('fecha_fin'),
        'tipo': request.form.get('tipo'),
        'observaciones': request.form.get('observaciones')
    }
    return redirect(url_for('mis_ausencias'))

@app.route('/control-horario')
@login_required
def control_horario():
    hoy = date.today()

    # Obtener el fichaje del día si existe
    fichaje = Fichaje.query.filter_by(usuario_id=current_user.id, fecha=hoy, eliminado=False).first()

    registros_hoy = []
    if fichaje:
        registros_hoy = RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False).order_by(RegistroHorario.timestamp).all()

    # Preparar variables
    estado = "Fuera"
    ultima_entrada = None
    ultima_salida = None
    total_segundos = 0
    tramos = []

    # Calcular estado y tramos
    i = 0
    while i < len(registros_hoy):
        if registros_hoy[i].tipo == 'entrada':
            entrada = registros_hoy[i]
            ultima_entrada = entrada.timestamp

            if i + 1 < len(registros_hoy) and registros_hoy[i + 1].tipo == 'salida':
                salida = registros_hoy[i + 1]
                ultima_salida = salida.timestamp
                diff = (salida.timestamp - entrada.timestamp).total_seconds()
                if diff > 0:
                    total_segundos += diff
                    tramos.append({
                        'inicio': entrada.timestamp.strftime('%H:%M'),
                        'fin': salida.timestamp.strftime('%H:%M'),
                        'duracion': '%02d:%02d' % (diff // 3600, (diff % 3600) // 60)
                    })
                i += 2
            else:
                # Entrada sin salida aún
                estado = "Dentro"
                i += 1
        elif registros_hoy[i].tipo == 'salida':
            ultima_salida = registros_hoy[i].timestamp
            estado = "Fuera"
            i += 1
        else:
            i += 1

    # Formateo de tiempo total
    horas = int(total_segundos // 3600)
    minutos = int((total_segundos % 3600) // 60)

    return render_template('control_horario.html',
                           fichaje=fichaje,
                           registros_hoy=registros_hoy,
                           hoy=hoy,
                           estado=estado,
                           ultima_entrada=ultima_entrada,
                           ultima_salida=ultima_salida,
                           horas=horas,
                           minutos=minutos,
                           tramos=tramos)

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    hoy = date.today()
    primer_dia = hoy.replace(day=1)
    dia = primer_dia

    #puestos_trabajan_sabados = ["Comercial", "Mostrador", "Taller", "Mozo Almacén"]
    mostrar_sabados = True

    sabados_mes = []
    while dia.month == hoy.month:
        if dia.weekday() == 5:
            sabados_mes.append(dia)
        dia += timedelta(days=1)

    asignados = {s.fecha for s in SabadoAsignado.query.filter_by(usuario_id=current_user.id).filter(
        SabadoAsignado.fecha.between(primer_dia, dia - timedelta(days=1))
    ).all()}

    if request.method == 'POST' and mostrar_sabados:
        seleccionados = request.form.getlist('sabados[]')
        SabadoAsignado.query.filter_by(usuario_id=current_user.id).filter(
            SabadoAsignado.fecha.between(primer_dia, dia - timedelta(days=1))
        ).delete()
        for fecha_str in seleccionados:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            db.session.add(SabadoAsignado(usuario_id=current_user.id, fecha=fecha))
        db.session.commit()
        flash("Tus sábados asignados se han guardado correctamente.")
        return redirect(url_for('perfil'))

    return render_template('perfil.html',
                           mostrar_sabados=mostrar_sabados,
                           sabados=sabados_mes,
                           asignados=asignados)

@app.route('/api/debo_fichar')
@login_required
def api_debo_fichar():
    hoy = date.today()
    ahora = datetime.now(zona_es).time()
    dia_semana = hoy.weekday()  # 0=lunes, ..., 5=sábado, 6=domingo
    mes = hoy.month  # agosto = 8

    # En agosto, nunca se trabaja los sábados
    if dia_semana == 5 and mes == 8:
        return {'debo_fichar': False}

    # Si hoy es sábado, solo se debe fichar si el usuario está asignado a ese sábado
    if dia_semana == 5:
        asignado = SabadoAsignado.query.filter_by(usuario_id=current_user.id, fecha=hoy).first()
        if not asignado:
            return {'debo_fichar': False}

    # Si el usuario tiene una ausencia hoy, no debe fichar
    if Ausencia.query.filter_by(usuario_id=current_user.id, fecha=hoy).first():
        return {'debo_fichar': False}

    # Obtener registros de hoy
    fichaje = Fichaje.query.filter_by(usuario_id=current_user.id, fecha=hoy, eliminado=False).first()
    registros = []
    if fichaje:
        registros = RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False).all()

    # Comprobar si ya ha fichado por la mañana o por la tarde
    ha_fichado_m = any(r.timestamp.time() <= datetime.strptime("14:00", "%H:%M").time() for r in registros)
    ha_fichado_t = any(r.timestamp.time() >= datetime.strptime("16:00", "%H:%M").time() for r in registros)

    # Franja crítica: 10:20–10:30 y 16:20–16:30
    debo_m = datetime.strptime("10:20", "%H:%M").time() <= ahora <= datetime.strptime("10:30", "%H:%M").time()
    debo_t = datetime.strptime("16:20", "%H:%M").time() <= ahora <= datetime.strptime("16:30", "%H:%M").time()

    # En agosto (días laborales), solo se controla la mañana
    if mes == 8:
        if debo_m and not ha_fichado_m:
            return {'debo_fichar': True}
    else:
        if (debo_m and not ha_fichado_m) or (debo_t and not ha_fichado_t):
            return {'debo_fichar': True}

    return {'debo_fichar': False}

@app.route('/admin/ausencias/<int:usuario_id>')
@login_required
def admin_ausencias_usuario(usuario_id):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    usuario = Usuario.query.get_or_404(usuario_id)
    hoy = date.today()

    ausencias_query = Ausencia.query.filter_by(usuario_id=usuario_id).order_by(Ausencia.fecha).all()

    datos = []
    total_utilizados = 0
    total_solicitados = 0

    for a in ausencias_query:
        dias = 1  # Cada ausencia es 1 día (puedes adaptarlo si manejas rangos)
        if a.tipo == "Vacaciones":
            if a.fecha < hoy:
                total_utilizados += dias
            else:
                total_solicitados += dias

        datos.append({
            'fecha': a.fecha.strftime('%d/%m/%Y'),
            'fecha_iso': a.fecha.isoformat(),
            'tipo': a.tipo,
            'observaciones': a.observaciones or '',
            'creado_por_admin': a.creado_por_admin
        })

    total_vacaciones = 30
    saldo_disponible = total_vacaciones - total_utilizados - total_solicitados

    return render_template(
        'admin_ausencias_usuario.html',
        usuario=usuario,
        ausencias=datos,
        saldo_disponible=saldo_disponible,
        saldo_solicitado=total_solicitados,
        saldo_utilizado=total_utilizados
    )

@app.route('/admin/registrar-ausencia/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
def admin_registrar_ausencia(usuario_id):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    usuario = Usuario.query.get_or_404(usuario_id)

    if request.method == 'POST':
        desde_str = request.form.get('fecha_desde')
        hasta_str = request.form.get('fecha_hasta')
        tipo = request.form.get('tipo')
        observaciones = request.form.get('observaciones', '').strip()

        try:
            fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()

            if fecha_hasta < fecha_desde:
                flash('La fecha "Hasta" no puede ser anterior a "Desde".', 'danger')
                return redirect(request.url)

            dias_ausencia = [
                fecha_desde + timedelta(days=i)
                for i in range((fecha_hasta - fecha_desde).days + 1)
            ]

            creadas = 0
            for dia in dias_ausencia:
                ya_existe = Ausencia.query.filter_by(usuario_id=usuario.id, fecha=dia).first()
                if not ya_existe:
                    nueva = Ausencia(
                        usuario_id=usuario.id,
                        fecha=dia,
                        tipo=tipo,
                        observaciones=observaciones,
                        creado_por_admin=True
                    )
                    db.session.add(nueva)
                    creadas += 1

            db.session.commit()

            if creadas:
                flash(f'{creadas} día(s) de ausencia registrados para {usuario.nombre}.', 'success')
            else:
                flash('No se registraron nuevas ausencias porque ya existían.', 'warning')

            return redirect(url_for('admin_usuarios'))

        except ValueError:
            flash('Fechas inválidas.', 'danger')
            return redirect(request.url)

    return render_template('admin_registrar_ausencia.html', usuario=usuario)

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    hoy = date.today()
    EXCLUIDOS_DASHBOARD = [7, 15]
    usuarios = Usuario.query.filter(~Usuario.id.in_(EXCLUIDOS_DASHBOARD)).order_by(Usuario.nombre).all()
    datos = []

    for user in usuarios:
        fichaje = Fichaje.query.filter_by(usuario_id=user.id, fecha=hoy, eliminado=False).first()
        registros = []
        total = timedelta()
        entrada_m = salida_m = entrada_t = salida_t = ''
        estado = '⛔ Sin fichaje'  # Valor por defecto

        if fichaje:
            registros = RegistroHorario.query.filter_by(fichaje_id=fichaje.id, eliminado=False)\
                .order_by(RegistroHorario.timestamp).all()
            num_registros = len(registros)

            if num_registros > 0:
                if registros[-1].tipo == 'entrada':
                    estado = '✅ Dentro'
                elif registros[-1].tipo == 'salida':
                    estado = '❌ Fuera'

                # Entrada mañana
                if num_registros >= 1:
                    entrada_m = registros[0].timestamp.strftime('%H:%M')
                # Salida mañana
                if num_registros >= 2:
                    salida_m = registros[1].timestamp.strftime('%H:%M')
                    total += registros[1].timestamp - registros[0].timestamp
                # Entrada tarde
                if num_registros >= 3:
                    entrada_t = registros[2].timestamp.strftime('%H:%M')
                # Salida tarde
                if num_registros >= 4:
                    salida_t = registros[3].timestamp.strftime('%H:%M')
                    total += registros[3].timestamp - registros[2].timestamp

        ausencia = Ausencia.query.filter_by(usuario_id=user.id, fecha=hoy).first()
        tipo_ausencia = ausencia.tipo if ausencia else None

        datos.append({
            'usuario': user.nombre,
            'estado': estado,
            'entrada_m': entrada_m,
            'salida_m': salida_m,
            'entrada_t': entrada_t,
            'salida_t': salida_t,
            'total': str(total) if total.total_seconds() > 0 else ('En curso' if estado == '✅ Dentro' else '-'),
            'ausencia': tipo_ausencia or '-'
        })

    return render_template('admin_dashboard.html', datos=datos, hoy=hoy)

@app.route('/cambiar-contrasena', methods=['POST'])
@login_required
def cambiar_contrasena():
    nueva = request.form.get('nueva_contrasena')
    confirmar = request.form.get('confirmar_contrasena')

    if not nueva or not confirmar:
        flash("Debes completar ambos campos.", "danger")
        return redirect(url_for('perfil'))

    if nueva != confirmar:
        flash("Las contraseñas no coinciden.", "danger")
        return redirect(url_for('perfil'))

    if len(nueva) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "danger")
        return redirect(url_for('perfil'))

    try:
        current_user.password_hash = generate_password_hash(nueva)
        db.session.commit()
        flash("Contraseña actualizada correctamente.", "success")
    except Exception as e:
        flash("Error al actualizar la contraseña.", "danger")

    return redirect(url_for('perfil'))

@app.route('/admin/crear-fichaje', methods=['GET', 'POST'])
@app.route('/admin/crear-fichaje/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
def crear_fichaje_admin(usuario_id=None):
    if not current_user.admin:
        flash("Acceso denegado.")
        return redirect(url_for('fichar'))

    usuarios = Usuario.query.order_by(Usuario.nombre).all()

    if request.method == 'POST':
        usuario_id = int(request.form.get('usuario_id'))
        fecha_str = request.form.get('fecha')
        entrada_str = request.form.get('entrada')
        salida_str = request.form.get('salida')

        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            hora_entrada = datetime.strptime(entrada_str, '%H:%M').time()
            hora_salida = datetime.strptime(salida_str, '%H:%M').time()

            if hora_entrada >= hora_salida:
                flash("La hora de entrada debe ser anterior a la de salida.", "danger")
                return redirect(request.url)

            ya_existe = Fichaje.query.filter_by(usuario_id=usuario_id, fecha=fecha).first()
            if ya_existe:
                flash("Ya existe un fichaje para ese usuario en esa fecha.", "warning")
                return redirect(url_for('editar_fichaje_admin', fichaje_id=ya_existe.id))

            # Crear nuevo fichaje y registros
            fichaje = Fichaje(
                usuario_id=usuario_id,
                fecha=fecha,
                fecha_creacion=datetime.now(),
                creado_por_admin=True,
                eliminado=False
            )
            db.session.add(fichaje)
            db.session.flush()

            r1 = RegistroHorario(
                fichaje_id=fichaje.id,
                tipo='entrada',
                timestamp=datetime.combine(fecha, hora_entrada),
                creado_por_admin=True,
                eliminado=False
            )
            r2 = RegistroHorario(
                fichaje_id=fichaje.id,
                tipo='salida',
                timestamp=datetime.combine(fecha, hora_salida),
                creado_por_admin=True,
                eliminado=False
            )
            db.session.add_all([r1, r2])
            db.session.commit()

            flash("Fichaje creado correctamente.", "success")
            return redirect(url_for('admin_fichajes_usuario', usuario_id=usuario_id))

        except Exception as e:
            flash("Error al crear el fichaje.", "danger")

    return render_template('admin_crear_fichaje.html', usuarios=usuarios, usuario_id_preseleccionado=usuario_id)