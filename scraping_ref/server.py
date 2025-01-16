from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from io import BytesIO
import openpyxl
import json
import logging

# Configuración básica de logs
logging.basicConfig(
    filename="\scraping_ref\\logs\\server.log",  # Cambia esta ruta si es necesario
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

app = FastAPI()

# Configuración de CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/buscar")
async def buscar_referencia(reference: str):
    """
    Endpoint para buscar referencias con resultados parciales usando GET.
    """
    logging.info(f"Petición GET /buscar con referencia: {reference}")
    if not reference:
        logging.warning("Referencia no proporcionada.")
        raise HTTPException(status_code=400, detail="Referencia no proporcionada.")

    try:
        from scraping import buscar_referencia_parcial  # Importar función generadora

        async def stream_results():
            for result in buscar_referencia_parcial(reference):
                logging.debug(f"Resultado encontrado: {result}")
                yield f"data: {json.dumps(result)}\n\n"
            yield "data: {\"completed\": true}\n\n"

        return StreamingResponse(stream_results(), media_type="text/event-stream")
    except Exception as e:
        logging.error(f"Error en '/buscar': {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor.")

@app.get("/")
async def root():
    return {"message": "El servidor está funcionando correctamente"}


@app.post("/descargar")
async def descargar_resultados(request: Request):
    """
    Endpoint para descargar los resultados en Excel o PDF.
    """
    try:
        data = await request.json()
        results = data.get('results')
        file_type = data.get('file_type', 'excel')
        searched_reference = data.get('reference', "Referencia no proporcionada")

        logging.info(f"Petición POST /descargar con formato {file_type} y referencia {searched_reference}")

        if not results:
            logging.warning("No se proporcionaron resultados para descargar.")
            return JSONResponse(content={"success": False, "error": "No se proporcionaron resultados para descargar."}, status_code=400)

        if file_type == 'excel':
            output = BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Resultados"

            # Escribir encabezados
            sheet.append([
                "Referencia Buscada", "Nombre", "PVP", "Descuento",
                "Precio Final", "Descripción", "Imagen URL",
                "Referencia Competencia", "Competencia"
            ])

            # Escribir resultados
            for item in results:
                sheet.append([
                    searched_reference,
                    item.get('name', ""),
                    item.get('pvp', ""),
                    item.get('discount', ""),
                    item.get('final_price', ""),
                    item.get('description', ""),
                    item.get('image_url', ""),
                    item.get('ref_competencia', ""),
                    item.get('competencia', "")
                ])

            workbook.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=resultados.xlsx"}
            )

        elif file_type == 'pdf':
            from fpdf import FPDF
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("Arial", size=12)

            # Escribir encabezados
            pdf.set_font("Arial", style="B", size=12)
            pdf.cell(0, 10, "Resultados de la búsqueda", ln=True, align="C")
            pdf.ln(10)

            pdf.set_font("Arial", size=10)
            for item in results:
                pdf.cell(0, 10, f"Referencia Buscada: {searched_reference}", ln=True)
                pdf.cell(0, 10, f"Nombre: {item.get('name', '')}", ln=True)
                pdf.cell(0, 10, f"PVP: {item.get('pvp', '')}", ln=True)
                pdf.cell(0, 10, f"Descuento: {item.get('discount', '')}", ln=True)
                pdf.cell(0, 10, f"Precio Final: {item.get('final_price', '')}", ln=True)
                pdf.cell(0, 10, f"Descripción: {item.get('description', '')}", ln=True)
                pdf.cell(0, 10, f"Referencia Competencia: {item.get('ref_competencia', '')}", ln=True)
                pdf.cell(0, 10, f"Competencia: {item.get('competencia', '')}", ln=True)
                pdf.ln(5)

            output = BytesIO()
            pdf.output(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=resultados.pdf"}
            )
        else:
            logging.warning(f"Formato no soportado: {file_type}")
            return JSONResponse(content={"success": False, "error": "Tipo de archivo no soportado."}, status_code=400)

    except Exception as e:
        logging.error(f"Error en '/descargar': {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor.")

# Ejecutar el servidor con Uvicorn
if __name__ == "__main__":
    import uvicorn
    logging.info("Iniciando el servidor FastAPI con Uvicorn.")
    uvicorn.run("server:app", host="0.0.0.0", port=5001, log_level="info")
